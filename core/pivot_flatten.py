"""Aplana tablas dinámicas de Excel (de cualquier layout: Compacto, Esquema
o Tabular; con o sin celdas combinadas; con o sin subtotales/total general)
para que el resto del pipeline (loader → cleaner → schema → profile) las
trate como una tabla plana normal.

Por qué hacía falta: pd.read_excel(header=None) lee la cuadrícula tal cual
quedó guardada en el Excel, y una tabla dinámica rompe 3 supuestos que el
resto del código sí hace sobre una tabla "normal":

1. Encabezado de una sola fila. Una dinámica con 2+ campos de columna
   (p. ej. Año arriba, Trimestre debajo) tiene 2+ filas de encabezado —
   _read_excel_sheet (ui/../core/loader.py) solo sabía elegir UNA.
2. Sin celdas vacías "heredadas". En layout Compacto/Esquema, la etiqueta
   de un grupo de filas (p. ej. "Región") solo aparece en la primera fila
   del grupo — el resto llegan vacías (por celda combinada real, o porque
   Excel simplemente no repite el texto). Sin relleno hacia abajo, esas
   filas quedan sin categoría.
3. Sin filas de subtotal/total mezcladas con los datos. "Total Norte",
   "Subtotal", "Total general"... son filas de AGREGADO, no un registro
   más — si se cuentan como dato, las sumas/promedios quedan infladas.

Cada paso de este módulo es seguro sobre una tabla plana normal: si no
encuentra nada que aplanar, no cambia nada (ver los tests al final del
archivo, ejecutables con `python -m core.pivot_flatten`).
"""
from __future__ import annotations

import io
import re

import pandas as pd

# Coincidencia EXACTA (toda la celda, no una palabra suelta dentro de un
# nombre) — así "Total Play" (un cliente real) o "Totalizadores S.A." nunca
# se confunden con una fila de agregado.
_TOTAL_EXACT_RE = re.compile(
    r"^(total\s*general|gran\s*total|grand\s*total|totales?|sub\s*-?\s*totales?)$",
    re.I,
)
# Coincidencia de PREFIJO ("Total Norte", "Total Región X" — el subtotal por
# grupo que arma Excel al activar "Subtotales" en una dinámica). Se usa con
# una condición extra (ver total_row_mask) para no atrapar categorías reales
# que solo empiezan con esa palabra.
_TOTAL_PREFIX_RE = re.compile(r"^(total|sub\s*-?\s*total)\s+\S", re.I)


def _cell_text(v) -> str:
    if pd.isna(v):
        return ""
    return re.sub(r"\s+", " ", str(v)).strip()


def _xlsx_merged_ranges_by_sheet(data: bytes) -> dict:
    """{nombre_de_hoja: [(r1,c1,r2,c2), ...]} de TODO el .xlsx/.xlsm en una
    sola pasada ligera sobre el XML interno (es un ZIP) — nunca carga el
    árbol completo de celdas de openpyxl.load_workbook().

    Por qué el cambio: la primera versión llamaba a
    openpyxl.load_workbook(...) — que sí carga cada celda de cada hoja en
    memoria como objetos — UNA VEZ POR HOJA (load_workbook() se invocaba
    dentro de _read_excel_sheet, que corre por hoja). Un Excel de varias
    hojas terminaba parseándose por completo tantas veces como hojas
    tuviera, encima de lo que pandas ya hace — en Streamlit Cloud (memoria
    limitada) eso se tradujo en que el proceso se quedaba sin memoria o
    colgado al subir un archivo, y el healthcheck lo mataba
    ("connection reset by peer", no un error de Python normal).

    Esto en cambio solo lee <mergeCells> del XML de cada hoja (texto plano,
    sin instanciar ninguna celda) y se llama UNA sola vez por archivo
    completo, sin importar cuántas hojas tenga."""
    import zipfile
    from xml.etree import ElementTree as ET
    from openpyxl.utils.cell import range_boundaries

    ns_main = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
    ns_rel = "http://schemas.openxmlformats.org/officeDocument/2006/relationships"
    ns_pkg_rel = "http://schemas.openxmlformats.org/package/2006/relationships"
    try:
        with zipfile.ZipFile(io.BytesIO(data)) as z:
            names = set(z.namelist())
            wb_xml = ET.fromstring(z.read("xl/workbook.xml"))
            rid_by_name = {
                sh.get("name"): sh.get(f"{{{ns_rel}}}id")
                for sh in wb_xml.findall(f"{{{ns_main}}}sheets/{{{ns_main}}}sheet")
            }
            wb_rels = ET.fromstring(z.read("xl/_rels/workbook.xml.rels"))
            target_by_rid = {
                rel.get("Id"): rel.get("Target")
                for rel in wb_rels.findall(f"{{{ns_pkg_rel}}}Relationship")
            }
            out = {}
            for sheet_name, rid in rid_by_name.items():
                target = target_by_rid.get(rid) or ""
                if not target:
                    continue
                # El Target de un Relationship en OOXML es o bien "absoluto
                # al paquete" (empieza con "/", p. ej. "/xl/worksheets/
                # sheet1.xml" — algunos generadores de Excel lo escriben
                # así) o relativo a la carpeta del propio .rels ("xl/_rels/",
                # así que relativo a "xl/", p. ej. "worksheets/sheet1.xml").
                # Tratar SIEMPRE el segundo caso sin distinguir el primero
                # producía "xl/xl/worksheets/sheet1.xml" (duplicado, no
                # existe en el zip) cada vez que el Target ya venía con "/"
                # — por eso nunca se encontraban celdas combinadas.
                sheet_path = target.lstrip("/") if target.startswith("/") else f"xl/{target}"
                if sheet_path not in names:
                    continue
                sheet_xml = ET.fromstring(z.read(sheet_path))
                ranges = []
                for mc in sheet_xml.findall(f"{{{ns_main}}}mergeCells/{{{ns_main}}}mergeCell"):
                    ref = mc.get("ref")
                    if not ref:
                        continue
                    try:
                        c1, r1, c2, r2 = range_boundaries(ref)
                    except ValueError:
                        continue
                    ranges.append((r1 - 1, c1 - 1, r2 - 1, c2 - 1))
                out[sheet_name] = ranges
            return out
    except Exception:
        return {}


def _xls_merged_ranges_by_sheet(data: bytes) -> dict:
    """Igual que la anterior pero para .xls legado (vía xlrd) — un solo
    xlrd.open_workbook() para todo el archivo, reutilizado por cada hoja en
    vez de reabrirlo por cada una."""
    try:
        import xlrd
        book = xlrd.open_workbook(file_contents=data)
        out = {}
        for sheet in book.sheets():
            out[sheet.name] = [(r0, c0, r1 - 1, c1 - 1) for (r0, r1, c0, c1) in sheet.merged_cells]
        return out
    except Exception:
        return {}


def merged_ranges_by_sheet(data: bytes, filename: str) -> dict:
    """{nombre_de_hoja: [(r1,c1,r2,c2), ...]} para TODO el archivo, en una
    sola pasada — se llama una vez por archivo (ver core/loader.py), nunca
    por hoja. {} si el formato no expone celdas combinadas (.xlsb) o si algo
    falla — nunca debe tumbar la carga del archivo."""
    name = filename.lower()
    if name.endswith((".xlsx", ".xlsm")):
        return _xlsx_merged_ranges_by_sheet(data)
    if name.endswith(".xls"):
        return _xls_merged_ranges_by_sheet(data)
    return {}


def fill_merged_cells(raw: pd.DataFrame, ranges) -> int:
    """Rellena, en el propio raw (in place), cada rango combinado con el
    valor de su celda superior-izquierda — la única que Excel guarda
    realmente; el resto de la combinación llega vacía. Devuelve cuántas
    celdas se rellenaron, para el log de calidad.

    Antes de escribir, sube a dtype `object` las columnas que participan en
    algún rango combinado (nunca la hoja entera — sería carísimo en un
    archivo grande sin necesidad). Por qué hace falta: `raw` es la
    cuadrícula cruda, tipada columna por columna por pandas según lo que
    haya en ESA columna a lo largo de TODA la hoja — una columna que en
    todas las demás filas es numérica queda tipada como float/entero
    "nulificable" (Float64/Int64). El título de un reporte, combinado en la
    fila de arriba, puede caer justo en una columna así — y escribir texto
    ahí truena ("Invalid value '...' for dtype 'Float64'") porque esos
    dtypes NO aceptan texto, a diferencia del float64 normal de numpy.
    `object` acepta cualquier tipo, así que el resto del pipeline (que ya
    maneja columnas mixtas más abajo, en cleaner.py) sigue funcionando
    igual, solo que sin ese riesgo de tipo."""
    if not ranges:
        return 0
    n_rows, n_cols = len(raw.index), len(raw.columns)
    merge_cols = {c for (_, c1, _, c2) in ranges for c in range(c1, c2 + 1) if c < n_cols}
    for c in merge_cols:
        if raw.dtypes.iloc[c] != object:
            raw.isetitem(c, raw.iloc[:, c].astype(object))
    filled = 0
    for (r1, c1, r2, c2) in ranges:
        if r1 >= n_rows or c1 >= n_cols:
            continue
        try:
            value = raw.iat[r1, c1]
        except Exception:
            continue
        if pd.isna(value):
            continue
        for r in range(r1, min(r2, n_rows - 1) + 1):
            for c in range(c1, min(c2, n_cols - 1) + 1):
                if r == r1 and c == c1:
                    continue
                if pd.isna(raw.iat[r, c]):
                    raw.iat[r, c] = value
                    filled += 1
    return filled


def is_super_header_row(values: list[str]) -> bool:
    """True si esta fila agrupa a la de abajo (p. ej. "2024" repetido sobre
    "Q1 Q2 Q3 Q4") en vez de ser un título suelto o ya la fila de nombres de
    columna real. Repetición (distintos < no-vacíos) es la señal: un título
    ocupa TODO el ancho de la fila tras el relleno de combinadas (título);
    un encabezado normal no repite nada (distintos==no vacíos). Un solo
    valor repetido que NO ocupa todo el ancho (p. ej. "2024" en B:E con la
    columna A —la de la etiqueta de fila— en blanco) sí cuenta: es un
    único grupo de nivel superior, no un título."""
    nonempty = [v for v in values if v]
    if len(nonempty) < 2:
        return False
    distinct = len(set(nonempty))
    if distinct == len(nonempty):
        return False
    if distinct <= 1 and len(nonempty) == len(values):
        return False
    string_ratio = sum(not re.fullmatch(r"[-+]?\d+(?:[.,]\d+)?", v) for v in nonempty) / len(nonempty)
    # Agrupar por año ("2024" repetido sobre varias columnas) es tan común
    # como agrupar por texto, pero "2024" es numérico → string_ratio solo no
    # lo detectaría. Una repetición FUERTE (cada valor cubre en promedio 2+
    # columnas) es igual de buena señal de agrupación aunque sea numérica.
    strong_repetition = distinct <= max(1, len(nonempty) // 2)
    return string_ratio >= 0.5 or strong_repetition


def combine_header_rows(top: list[str], bottom: list[str]) -> list[str]:
    """Unifica una fila super-encabezado (agrupa columnas; puede traer
    huecos si el agrupado no vino de una celda combinada real) con la fila
    de encabezado real debajo, tipo "2024 · Q1"."""
    combined = []
    last_top = ""
    for top_v, bottom_v in zip(top, bottom):
        if top_v:
            last_top = top_v
        if last_top and bottom_v:
            combined.append(f"{last_top} · {bottom_v}")
        else:
            combined.append(bottom_v or last_top)
    return combined


def total_row_mask(df: pd.DataFrame, label_cols) -> pd.Series:
    """True en cada fila de subtotal/total general. Dos criterios,
    deliberadamente conservadores para no tocar datos reales:
    - Coincidencia EXACTA de la primera celda no vacía de la fila
      ("Total", "Subtotal", "Total general"...) → siempre se excluye.
    - Coincidencia de PREFIJO ("Total Norte") en una columna de etiqueta,
      solo si ADEMÁS el resto de columnas de etiqueta de esa fila están
      vacías — la firma real de una fila de rollup (agrupa, no describe un
      registro). Evita que una categoría real que arranca con "Total..."
      (p. ej. la empresa "Total Play") se confunda con un agregado, porque
      esa sí trae sus demás columnas de etiqueta normalmente llenas.
    """
    label_cols = [c for c in label_cols if c in df.columns]

    def _row_is_total(row) -> bool:
        first_val, first_col = None, None
        for c in df.columns:
            s = _cell_text(row[c])
            if s:
                first_val, first_col = s, c
                break
        if not first_val:
            return False
        if _TOTAL_EXACT_RE.match(first_val):
            return True
        if _TOTAL_PREFIX_RE.match(first_val) and first_col in label_cols:
            others = [c for c in label_cols if c != first_col]
            if others and all(not _cell_text(row[c]) for c in others):
                return True
        return False

    return df.apply(_row_is_total, axis=1)


def staircase_fill(df: pd.DataFrame, columns) -> int:
    """Relleno hacia abajo (ffill), pero solo dentro de columnas que ya se
    confirmó que son de etiqueta/agrupación de una dinámica (ver el gateo en
    _read_excel_sheet — nunca se llama a esto sobre una tabla que no mostró
    ninguna otra señal de ser dinámica). Nunca rellena antes del primer
    valor real de la columna (si la propia tabla empieza sin ese dato, se
    queda sin dato, no se inventa uno)."""
    filled = 0
    for c in columns:
        if c not in df.columns:
            continue
        before = df[c].copy()
        df[c] = df[c].ffill()
        # No rellenar filas que nunca tuvieron un valor real arriba.
        first_valid = before.first_valid_index()
        if first_valid is not None:
            mask_before_first = df.index < first_valid
            df.loc[mask_before_first, c] = before.loc[mask_before_first]
        filled += int((before.isna() & df[c].notna()).sum())
    return filled


def flatten_pivot_grid(raw: pd.DataFrame, header_score_fn, norm_header_fn, make_unique_fn):
    """Punto de entrada único, usado por core/loader.py en vez del antiguo
    "una sola fila de encabezado, listo". Devuelve (data_df, log) — log es
    una lista de mensajes en español, listos para sumarse al log de
    limpieza que ya se muestra en la pestaña Calidad.

    header_score_fn/norm_header_fn/make_unique_fn: las mismas funciones de
    loader.py (_header_score, _norm_header, _make_unique_columns) — se
    reciben por parámetro en vez de importarlas para no crear un ciclo de
    imports entre los dos módulos.
    """
    log: list[str] = []
    if raw.empty:
        return raw, log

    limit = min(len(raw), 15)
    candidates = [(i, header_score_fn(raw.iloc[i])) for i in range(limit)]
    best_i, best_score = max(candidates, key=lambda x: x[1])
    first_score = candidates[0][1]
    confident = True
    if first_score >= 0.58 and first_score >= best_score - 0.04:
        header_i = 0
    elif best_score >= 0.55:
        header_i = best_i
    else:
        header_i = 0
        confident = False

    def _nonempty_count(i):
        return sum(1 for v in raw.iloc[i].tolist() if norm_header_fn(v))

    def _is_title_row(i):
        """Fila con UN solo valor repetido en TODAS sus celdas, sin ningún
        hueco (p. ej. el título de un reporte, "INFORME PDC TaT TROPAS",
        combinado en toda la fila y ya rellenado por fill_merged_cells).
        Nunca es un encabezado real — un encabezado real, por definición,
        nombra columnas DISTINTAS. Se comprueba aparte de is_super_header_row
        (que sí permite un valor repetido, pero solo si agrupa una PARTE del
        ancho — un super-encabezado real dos niveles abajo)."""
        values = [norm_header_fn(v) for v in raw.iloc[i].tolist()]
        nonempty = [v for v in values if v]
        return len(nonempty) == len(values) and len(nonempty) >= 2 and len(set(nonempty)) == 1

    # ── "Mirar abajo": dos motivos para desconfiar de la fila elegida y
    # preferir la de abajo como encabezado real:
    # 1. Es un título de fila completa (_is_title_row) — nunca es un
    #    encabezado, así que este caso se corrige SIEMPRE, sin importar el
    #    puntaje (un título con una frase que por casualidad toca una
    #    palabra del diccionario semántico podría incluso puntuar "confident").
    # 2. (Solo si 1 no aplicó) NINGUNA fila superó el umbral de confianza de
    #    arriba — columnas tipo "Q1"/"Q2" casi nunca traen ninguna palabra
    #    del diccionario semántico de _header_score, así que un
    #    super-encabezado real puede quedar por debajo de 0.55 aunque sea
    #    perfectamente válido — la fila elegida por defecto (0) puede ser
    #    ese super-encabezado, con el encabezado real justo debajo; se
    #    detecta comparando: la fila de abajo puntúa claramente mejor Y
    #    tiene más celdas llenas (agrupa menos que lo que agrupa).
    # Sin ninguno de los dos, header_i se hubiera quedado apuntando a una
    # fila que no es el encabezado real, y esa fila real se habría leído
    # como si fuera un dato más. ──
    # Fila(s) de título descartadas — a diferencia de un super-encabezado,
    # un título NUNCA se combina con el encabezado real (produciría columnas
    # como "INFORME PDC TaT TROPAS · 2024"), se tira sin más.
    discarded_title_rows = 0
    while header_i + 1 < limit and _is_title_row(header_i):
        header_i += 1
        discarded_title_rows += 1
        confident = False  # la fila ahora elegida todavía no pasó el chequeo de puntaje
    if discarded_title_rows:
        log.append(
            f"{discarded_title_rows} fila(s) de título (texto repetido en toda la fila, "
            "sin relación con los datos) descartada(s) del encabezado."
        )

    header_rows = [header_i]
    if not confident and header_i + 1 < limit:
        below_score = candidates[header_i + 1][1]
        chosen_nonempty = _nonempty_count(header_i)
        below_nonempty = _nonempty_count(header_i + 1)
        if (
            below_score >= 0.45
            and below_score >= candidates[header_i][1] + 0.15
            and chosen_nonempty > 0
            and below_nonempty > chosen_nonempty
        ):
            header_rows = [header_i, header_i + 1]

    # ── Encabezado de varias filas: se absorben hasta 2 filas extra por
    # ARRIBA del primer renglón del encabezado mientras cada una siga
    # viéndose como un super-encabezado real (agrupa, no repite un título
    # suelto) — cubre el caso contrario al de arriba: cuando header_i SÍ
    # cayó bien en la fila de nombres de columna real, y el super-encabezado
    # (con sus celdas ya rellenadas si venían de una combinación real de
    # Excel) está un renglón antes. ──
    probe = header_rows[0] - 1
    levels_absorbed = 0
    while probe >= 0 and levels_absorbed < 2:
        values = [norm_header_fn(v) for v in raw.iloc[probe].tolist()]
        if not is_super_header_row(values):
            break
        header_rows.insert(0, probe)
        probe -= 1
        levels_absorbed += 1

    if len(header_rows) > 1:
        combined = [norm_header_fn(v) for v in raw.iloc[header_rows[0]].tolist()]
        for r in header_rows[1:]:
            combined = combine_header_rows(combined, [norm_header_fn(v) for v in raw.iloc[r].tolist()])
        header_values = combined
        log.append(
            f"Encabezado de {len(header_rows)} filas combinadas en una sola "
            f"(filas {header_rows[0] + 1} a {header_rows[-1] + 1} del Excel)."
        )
    else:
        header_values = [norm_header_fn(v) for v in raw.iloc[header_i].tolist()]

    header = make_unique_fn(header_values)
    data_df = raw.iloc[header_rows[-1] + 1:].copy()
    data_df.columns = header
    data_df = data_df.dropna(axis=0, how="all").dropna(axis=1, how="all")
    data_df = data_df.reset_index(drop=True)
    if data_df.empty:
        return data_df, log

    # ── Filas de subtotal / total general ──
    label_cols = [
        c for c in data_df.columns
        if pd.to_numeric(data_df[c], errors="coerce").notna().mean() < 0.5
    ][:6]
    mask = total_row_mask(data_df, label_cols)
    n_total_rows = int(mask.sum())
    pivot_signal = n_total_rows > 0 or len(header_rows) > 1
    if n_total_rows:
        log.append(
            f"{n_total_rows} fila(s) de subtotal/total general excluida(s) "
            "automáticamente (no son un registro, son un agregado)."
        )
        data_df = data_df.loc[~mask].reset_index(drop=True)

    # ── Etiquetas de fila heredadas (layout Compacto/Esquema) ──
    # Gateado a propósito: solo se activa si YA hay otra señal de que esto
    # es una dinámica (encabezado multi-fila o alguna fila de total
    # detectada). Sin esa señal, una columna de texto con huecos se deja
    # tal cual — puede ser dato real ausente, no una etiqueta heredada.
    if pivot_signal:
        candidate_cols = []
        for c in label_cols[:4]:
            s = data_df[c]
            blank_ratio = s.isna().mean()
            if 0.03 <= blank_ratio <= 0.85 and pd.notna(s.iloc[0] if len(s) else None):
                candidate_cols.append(c)
        if candidate_cols:
            n_filled = staircase_fill(data_df, candidate_cols)
            if n_filled:
                cols_txt = ", ".join(str(c) for c in candidate_cols)
                log.append(
                    f"{n_filled} celda(s) de categoría heredadas de la fila de arriba "
                    f"(típico de tablas dinámicas), rellenadas en: {cols_txt}."
                )

    return data_df, log
