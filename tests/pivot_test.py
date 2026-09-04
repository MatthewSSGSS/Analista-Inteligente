"""Regresión del aplanador de tablas dinámicas (core/pivot_flatten.py).

Genera archivos .xlsx/.csv reales replicando las formas más comunes de
tabla dinámica de Excel (celdas combinadas, encabezado de varias filas,
subtotales, total general — en Compacto/Esquema, con y sin combinar) y
verifica que core/loader.py las aplane a una tabla normal. Incluye también
casos de control: datos planos normales, con o sin coincidencias que
podrían confundirse con una tabla dinámica, no deben tocarse.

Se corre igual que tests/smoke_test.py: PYTHONPATH=. python tests/pivot_test.py
"""
import io
import openpyxl

from core.loader import load_workbook


def _upload_xlsx(name, wb):
    b = io.BytesIO()
    wb.save(b)
    return type("Upload", (), {"getvalue": lambda self: b.getvalue(), "name": name})()


def _upload_csv(name, text):
    data = text.encode("utf-8")
    return type("Upload", (), {"getvalue": lambda self, d=data: d, "name": name})()


def check(label, condition):
    if not condition:
        raise AssertionError(label)
    print("OK  ", label)


# ── Caso 1: layout Compacto, CON celdas combinadas reales para "Región"
# (grupo de filas), encabezado de 1 fila, subtotal por región y total
# general al final. ──
def caso_1_merges_y_subtotales():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Ventas"
    ws.append(["Región", "Producto", "Ingresos"])
    for r in [
        ("Norte", "A", 100), ("Norte", "B", 150), ("Total Norte", "", 250),
        ("Sur", "A", 80), ("Sur", "B", 120), ("Sur", "C", 40), ("Total Sur", "", 240),
        ("Total general", "", 490),
    ]:
        ws.append(r)
    ws.merge_cells("A2:A3")
    ws.merge_cells("A5:A7")
    item = load_workbook(_upload_xlsx("dinamica_regiones.xlsx", wb))["sheets"]["Ventas"]
    df, log = item["processed"], item["profile"]["cleaning_log"]
    check("quita las 3 filas de subtotal/total", len(df) == 5)
    check("rellena la región heredada de la celda combinada", df["Región"].isna().sum() == 0)
    check("Norte aparece 2 veces, Sur 3 veces", (df["Región"] == "Norte").sum() == 2 and (df["Región"] == "Sur").sum() == 3)
    check("suma real (sin los totales) = 490", df["Ingresos"].sum() == 490)
    check("el log menciona celdas combinadas y subtotales", any("combinada" in x for x in log) and any("subtotal" in x.lower() for x in log))


# ── Caso 2: SIN celdas combinadas (texto simplemente en blanco), encabezado
# de 2 filas (Año arriba, Trimestre abajo), solo total general. ──
def caso_2_staircase_sin_merge_header_2_filas():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Trimestral"
    ws.append(["", "2024", "", "2025", ""])
    ws.append(["Vendedor", "Q1", "Q2", "Q1", "Q2"])
    ws.append(["Ana", 10, 12, 14, 16])
    ws.append(["", 11, 13, 15, 17])
    ws.append(["Beto", 20, 22, 24, 26])
    ws.append(["Total general", 41, 47, 53, 59])
    item = load_workbook(_upload_xlsx("dinamica_trimestral.xlsx", wb))["sheets"]["Trimestral"]
    df, log = item["processed"], item["profile"]["cleaning_log"]
    check("encabezado de 2 filas combinado en 1 (5 columnas)", len(df.columns) == 5 and "2024 · Q1" in df.columns)
    check("quita la fila de total general (2 personas quedan)", len(df) == 3)
    check("Ana hereda su nombre en la fila de abajo (sin merge real)", (df["Vendedor"] == "Ana").sum() == 2)
    check("el log menciona el encabezado combinado", any("Encabezado" in x for x in log))


# ── Caso 3: encabezado de 3 niveles (Año > Región > Métrica), con celdas
# combinadas en los 2 niveles superiores. ──
def caso_3_encabezado_3_niveles():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Multi"
    ws.append(["", "2024", "", "", ""])
    ws.append(["", "Norte", "", "Sur", ""])
    ws.append(["Vendedor", "Ingresos", "Unidades", "Ingresos", "Unidades"])
    ws.append(["Ana", 100, 10, 80, 8])
    ws.append(["Beto", 120, 12, 90, 9])
    ws.merge_cells("B1:E1")
    ws.merge_cells("B2:C2")
    ws.merge_cells("D2:E2")
    item = load_workbook(_upload_xlsx("multinivel.xlsx", wb))["sheets"]["Multi"]
    df = item["processed"]
    check("junta los 3 niveles del encabezado", any("2024" in c and "Norte" in c and "Ingresos" in c for c in df.columns))
    check("quedan las 2 filas de datos", len(df) == 2)


# ── Caso 4: CSV exportado de una dinámica (sin celdas combinadas posibles
# en CSV, pero con el mismo hueco de etiquetas + total general). ──
def caso_4_csv_con_forma_de_dinamica():
    csv_text = "Región,Producto,Ingresos\nNorte,A,100\n,B,150\nSur,A,80\n,B,120\nTotal general,,450\n"
    item = load_workbook(_upload_csv("export_dinamica.csv", csv_text))["sheets"]["CSV"]
    df = item["processed"]
    check("quita la fila de total general del CSV", len(df) == 4)
    check("Norte y Sur se heredan en la fila de abajo", (df["Región"] == "Norte").sum() == 2 and (df["Región"] == "Sur").sum() == 2)


# ── Control 1: tabla PLANA normal con una categoría real que EMPIEZA con
# "Total" (cliente real "Total Play") — no debe excluirse ni tocarse. ──
def control_1_categoria_real_total_play():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Clientes"
    ws.append(["Cliente", "Ciudad", "Ingresos"])
    ws.append(["Total Play", "Bogotá", 5000])
    ws.append(["Claro", "Medellín", 4000])
    ws.append(["Movistar", "Cali", 3000])
    item = load_workbook(_upload_xlsx("clientes.xlsx", wb))["sheets"]["Clientes"]
    df, log = item["processed"], item["profile"]["cleaning_log"]
    check("Total Play sigue en los datos", (df["Cliente"] == "Total Play").sum() == 1)
    check("siguen las 3 filas", len(df) == 3)
    check("no se detectó ningún subtotal (no debía aplicar)", not any("subtotal" in x.lower() for x in log))


# ── Control 2: columna de texto genuinamente dispersa (segundo apellido en
# blanco para varias personas) SIN ninguna otra señal de dinámica — no debe
# rellenarse con el valor de arriba. ──
def control_2_columna_dispersa_no_se_rellena():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Personas"
    ws.append(["Nombre", "Segundo apellido", "Edad"])
    ws.append(["Ana", "Gómez", 30])
    ws.append(["Luis", "", 40])
    ws.append(["Marta", "Ruiz", 25])
    ws.append(["Pedro", "", 35])
    item = load_workbook(_upload_xlsx("personas.xlsx", wb))["sheets"]["Personas"]
    df, log = item["processed"], item["profile"]["cleaning_log"]
    import pandas as pd
    check("el segundo apellido de Luis se queda vacío (no hereda 'Gómez')", pd.isna(df["Segundo apellido"].iloc[1]))
    check("el segundo apellido de Pedro se queda vacío (no hereda 'Ruiz')", pd.isna(df["Segundo apellido"].iloc[3]))
    check("no se activó ningún relleno heredado", not any("heredad" in x for x in log))


# ── Control 3: hoja de solo encabezado (sin filas de dato) no debe tronar
# el proceso de carga completo. ──
def control_3_hoja_sin_filas_de_dato():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "SoloEncabezado"
    ws.append(["Nombre", "Valor"])
    try:
        load_workbook(_upload_xlsx("solo_encabezado.xlsx", wb))
        check("hoja sin filas de dato no truena", True)
    except ValueError:
        check("hoja sin filas de dato: error controlado (esperado), no un crash", True)


if __name__ == "__main__":
    caso_1_merges_y_subtotales()
    caso_2_staircase_sin_merge_header_2_filas()
    caso_3_encabezado_3_niveles()
    caso_4_csv_con_forma_de_dinamica()
    control_1_categoria_real_total_play()
    control_2_columna_dispersa_no_se_rellena()
    control_3_hoja_sin_filas_de_dato()
    print("\nPivot test completado sin errores.")
